import argparse
import csv
import math
import random
from collections import defaultdict
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError as error:
    raise ImportError(
        "This script needs openpyxl. Install it once with: "
        "pip install openpyxl"
    ) from error


# =====================================================
# CONFIGURATION
# =====================================================

DEFAULT_INPUT = Path(
    r"C:\Users\Bouchra\Desktop\Self_Guided_Microsoft"
    r"\RAG_Project\evaluation_results.csv"
)

SAMPLE_SIZE = 30
RANDOM_SEED = 42
NOT_FOUND = "not found in the company documents"

LABELS = [
    "Supported",
    "Partially unsupported",
    "Hallucinated",
    "Unclear",
]


# =====================================================
# DATA HELPERS
# =====================================================

def clean(value):
    return str(value or "").strip()


def is_refusal(value):
    return NOT_FOUND in clean(value).lower()


def read_evaluation(path):
    with path.open("r", encoding="utf-8-sig", newline="") as file:
        reader = csv.DictReader(file)
        fieldnames = list(reader.fieldnames or [])
        rows = list(reader)

    required = {"Question", "Gold_Answer"}
    missing = required - set(fieldnames)

    if missing:
        raise ValueError(
            f"evaluation_results.csv is missing columns: {sorted(missing)}"
        )

    baseline_column = next(
        (
            name
            for name in ("Baseline", "Baseline_Answer")
            if name in fieldnames
        ),
        None,
    )
    rag_column = next(
        name for name in ("RAG", "RAG_Answer") if name in fieldnames
    )

    if baseline_column is None or rag_column is None:
        raise ValueError(
            "The file must contain Baseline and RAG answer columns."
        )

    if not rows:
        raise ValueError("evaluation_results.csv contains no rows.")

    return rows, baseline_column, rag_column


def allocate_sample(groups, sample_size):
    """Allocate a proportional sample while including every document."""
    documents = sorted(groups)
    total = sum(len(groups[document]) for document in documents)

    if sample_size < len(documents):
        raise ValueError(
            "The sample size must be at least the number of documents."
        )

    exact = {
        document: sample_size * len(groups[document]) / total
        for document in documents
    }
    allocation = {
        document: min(
            len(groups[document]),
            max(1, math.floor(exact[document])),
        )
        for document in documents
    }

    while sum(allocation.values()) < sample_size:
        candidates = [
            document
            for document in documents
            if allocation[document] < len(groups[document])
        ]
        selected = max(
            candidates,
            key=lambda document: exact[document] - allocation[document],
        )
        allocation[selected] += 1

    while sum(allocation.values()) > sample_size:
        candidates = [
            document
            for document in documents
            if allocation[document] > 1
        ]
        selected = max(
            candidates,
            key=lambda document: allocation[document] - exact[document],
        )
        allocation[selected] -= 1

    return allocation


def select_sample(rows, baseline_column, rag_column):
    """Select 30 questions answered by both Baseline and RAG."""
    eligible = []

    for row_number, row in enumerate(rows, start=2):
        baseline_answer = clean(row.get(baseline_column))
        rag_answer = clean(row.get(rag_column))
        question = clean(row.get("Question"))

        if not question or not baseline_answer or not rag_answer:
            continue

        # Refusals are coverage failures, not hallucinations.
        if is_refusal(baseline_answer) or is_refusal(rag_answer):
            continue

        copied = dict(row)
        copied["_Row"] = row_number
        eligible.append(copied)

    if len(eligible) < SAMPLE_SIZE:
        raise ValueError(
            f"Only {len(eligible)} jointly answered questions are available."
        )

    groups = defaultdict(list)
    for row in eligible:
        groups[clean(row.get("Document")) or "Not specified"].append(row)

    allocation = allocate_sample(groups, SAMPLE_SIZE)
    generator = random.Random(RANDOM_SEED)
    selected = []

    for document in sorted(groups):
        selected.extend(
            generator.sample(groups[document], allocation[document])
        )

    selected.sort(
        key=lambda row: (clean(row.get("Document")), int(row["_Row"]))
    )

    return selected, len(eligible)


# =====================================================
# EXCEL WORKBOOK
# =====================================================

def create_workbook(
    output_path,
    source_rows,
    sample_rows,
    eligible_count,
    baseline_column,
    rag_column,
):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Hallucination Evaluation"
    sheet.sheet_view.showGridLines = False

    # Colours and formatting.
    dark_blue = "17365D"
    blue = "2F75B5"
    light_blue = "D9EAF7"
    yellow = "FFF2CC"
    orange = "FCE4D6"
    green = "E2F0D9"
    white = "FFFFFF"
    dark_text = "1F2937"
    thin = Side(style="thin", color="B4C7DC")
    borders = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title and short instructions inside the same file.
    sheet.merge_cells("A1:J1")
    sheet["A1"] = "Hallucination Evaluation: Baseline LLM vs RAG"
    sheet["A1"].font = Font(bold=True, color=white, size=16)
    sheet["A1"].fill = PatternFill("solid", fgColor=dark_blue)
    sheet["A1"].alignment = Alignment(horizontal="center")

    sheet.merge_cells("A2:J2")
    sheet["A2"] = (
        "Fill only the two yellow label columns. "
        "The result table calculates automatically."
    )
    sheet["A2"].font = Font(italic=True, color=dark_text)
    sheet["A2"].fill = PatternFill("solid", fgColor=light_blue)
    sheet["A2"].alignment = Alignment(horizontal="center", wrap_text=True)

    sheet.merge_cells("A3:J3")
    sheet["A3"] = (
        "Labels: Supported | Partially unsupported | Hallucinated | "
        "Unclear. Verify claims against the named company document."
    )
    sheet["A3"].alignment = Alignment(wrap_text=True)

    sheet["A4"] = "Status"
    sheet["A4"].font = Font(bold=True, color=white)
    sheet["A4"].fill = PatternFill("solid", fgColor=dark_blue)
    sheet.merge_cells("B4:D4")
    sheet["B4"] = (
        '=IF(COUNTA($F$15:$F$44)+COUNTA($I$15:$I$44)=60,'
        '"Complete","Fill all 60 label cells")'
    )
    sheet["B4"].font = Font(bold=True, color="7F6000")
    sheet["B4"].fill = PatternFill("solid", fgColor=yellow)
    sheet["B4"].alignment = Alignment(horizontal="center")

    sheet.merge_cells("A5:J5")
    sheet["A5"] = (
        "Strict rate = Partially unsupported + Hallucinated. "
        "Severe rate = Hallucinated only."
    )
    sheet["A5"].fill = PatternFill("solid", fgColor=orange)
    sheet["A5"].alignment = Alignment(wrap_text=True)

    # Result table.
    sheet.merge_cells("A7:I7")
    sheet["A7"] = "RESULT TABLE"
    sheet["A7"].font = Font(bold=True, color=white, size=12)
    sheet["A7"].fill = PatternFill("solid", fgColor=blue)

    summary_headers = [
        "System",
        "Sampled",
        "Supported",
        "Partially unsupported",
        "Hallucinated",
        "Unclear",
        "Clear denominator",
        "Strict hallucination rate",
        "Severe hallucination rate",
    ]
    for column, header in enumerate(summary_headers, start=1):
        cell = sheet.cell(row=8, column=column, value=header)
        cell.font = Font(bold=True, color=dark_blue)
        cell.fill = PatternFill("solid", fgColor=light_blue)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = borders

    sheet["A9"] = "Baseline LLM"
    sheet["A10"] = "RAG"

    baseline_labels = "$F$15:$F$44"
    rag_labels = "$I$15:$I$44"

    for row_number, label_range in ((9, baseline_labels), (10, rag_labels)):
        sheet.cell(row=row_number, column=2, value="=COUNTA($A$15:$A$44)")
        sheet.cell(
            row=row_number,
            column=3,
            value=f'=COUNTIF({label_range},"Supported")',
        )
        sheet.cell(
            row=row_number,
            column=4,
            value=f'=COUNTIF({label_range},"Partially unsupported")',
        )
        sheet.cell(
            row=row_number,
            column=5,
            value=f'=COUNTIF({label_range},"Hallucinated")',
        )
        sheet.cell(
            row=row_number,
            column=6,
            value=f'=COUNTIF({label_range},"Unclear")',
        )
        sheet.cell(row=row_number, column=7, value=f"=SUM(C{row_number}:E{row_number})")
        sheet.cell(
            row=row_number,
            column=8,
            value=(
                f'=IF($B$4<>"Complete","",IF(G{row_number}=0,"",'
                f'(D{row_number}+E{row_number})/G{row_number}))'
            ),
        )
        sheet.cell(
            row=row_number,
            column=9,
            value=(
                f'=IF($B$4<>"Complete","",IF(G{row_number}=0,"",'
                f'E{row_number}/G{row_number}))'
            ),
        )

    for row in sheet.iter_rows(min_row=9, max_row=10, min_col=1, max_col=9):
        for cell in row:
            cell.border = borders
            cell.alignment = Alignment(horizontal="center", vertical="center")
        row[0].font = Font(bold=True, color=dark_blue)

    for cell in sheet["H"][8:10] + sheet["I"][8:10]:
        cell.number_format = "0.0%"

    summary_table = Table(displayName="HallucinationSummary", ref="A8:I10")
    summary_table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(summary_table)

    baseline_refusals = sum(
        is_refusal(row.get(baseline_column)) for row in source_rows
    )
    rag_refusals = sum(is_refusal(row.get(rag_column)) for row in source_rows)

    sheet.merge_cells("A11:J11")
    sheet["A11"] = (
        f"Coverage: Baseline answered {len(source_rows) - baseline_refusals}/"
        f"{len(source_rows)}; RAG answered {len(source_rows) - rag_refusals}/"
        f"{len(source_rows)} and refused {rag_refusals}. "
        "Refusals are not hallucinations."
    )
    sheet["A11"].fill = PatternFill("solid", fgColor=green)
    sheet["A11"].alignment = Alignment(wrap_text=True)

    # Annotation table in the same worksheet.
    sheet.merge_cells("A13:J13")
    sheet["A13"] = "ANNOTATION TABLE"
    sheet["A13"].font = Font(bold=True, color=white, size=12)
    sheet["A13"].fill = PatternFill("solid", fgColor=blue)

    annotation_headers = [
        "Sample ID",
        "Document",
        "Question",
        "Gold answer",
        "Baseline answer",
        "Baseline label",
        "Baseline evidence notes",
        "RAG answer",
        "RAG label",
        "RAG evidence notes",
    ]
    for column, header in enumerate(annotation_headers, start=1):
        cell = sheet.cell(row=14, column=column, value=header)
        cell.font = Font(bold=True, color=dark_blue)
        cell.fill = PatternFill("solid", fgColor=light_blue)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = borders

    for sample_number, row in enumerate(sample_rows, start=1):
        excel_row = 14 + sample_number
        values = [
            f"H{sample_number:03d}",
            clean(row.get("Document")),
            clean(row.get("Question")),
            clean(row.get("Gold_Answer")),
            clean(row.get(baseline_column)),
            "",
            "",
            clean(row.get(rag_column)),
            "",
            "",
        ]

        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row=excel_row, column=column, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin)

        sheet.cell(excel_row, 6).fill = PatternFill("solid", fgColor=yellow)
        sheet.cell(excel_row, 9).fill = PatternFill("solid", fgColor=yellow)

    label_validation = DataValidation(
        type="list",
        formula1='"Supported,Partially unsupported,Hallucinated,Unclear"',
        allow_blank=True,
    )
    sheet.add_data_validation(label_validation)
    label_validation.add("F15:F44")
    label_validation.add("I15:I44")

    fills = {
        "Supported": ("C6EFCE", "006100"),
        "Partially unsupported": ("FFEB9C", "9C6500"),
        "Hallucinated": ("FFC7CE", "9C0006"),
        "Unclear": ("E7E6E6", "595959"),
    }
    for label, (fill_colour, font_colour) in fills.items():
        for column in ("F", "I"):
            sheet.conditional_formatting.add(
                f"{column}15:{column}44",
                FormulaRule(
                    formula=[f'{column}15="{label}"'],
                    fill=PatternFill("solid", fgColor=fill_colour),
                    font=Font(color=font_colour),
                ),
            )

    annotation_table = Table(
        displayName="HallucinationAnnotations",
        ref="A14:J44",
    )
    annotation_table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(annotation_table)

    widths = {
        "A": 11,
        "B": 30,
        "C": 45,
        "D": 48,
        "E": 55,
        "F": 22,
        "G": 34,
        "H": 55,
        "I": 22,
        "J": 34,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    sheet.row_dimensions[1].height = 30
    sheet.row_dimensions[2].height = 32
    sheet.row_dimensions[3].height = 32
    sheet.row_dimensions[5].height = 32
    sheet.row_dimensions[11].height = 32
    for row_number in range(15, 45):
        sheet.row_dimensions[row_number].height = 84

    sheet.freeze_panes = "C15"
    sheet.auto_filter.ref = "A14:J44"

    # Ask Excel to recalculate the formulas when the file is opened.
    try:
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.calculation.calcMode = "auto"
    except AttributeError:
        pass

    workbook.save(output_path)


# =====================================================
# MAIN
# =====================================================

def parse_arguments():
    parser = argparse.ArgumentParser(
        description=(
            "Create one Excel file for manual hallucination evaluation. "
            "No Ollama or answer generation is used."
        )
    )
    parser.add_argument(
        "--input",
        type=Path,
        default=DEFAULT_INPUT,
        help="Path to evaluation_results.csv",
    )
    return parser.parse_args()


def main():
    args = parse_arguments()
    input_path = args.input.resolve()
    output_path = input_path.parent / "hallucination_evaluation.xlsx"

    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")

    if output_path.exists():
        raise FileExistsError(
            f"The output already exists and was not overwritten: {output_path}\n"
            "Rename it only if you intentionally want to create a new copy."
        )

    rows, baseline_column, rag_column = read_evaluation(input_path)
    sample_rows, eligible_count = select_sample(
        rows,
        baseline_column,
        rag_column,
    )

    create_workbook(
        output_path=output_path,
        source_rows=rows,
        sample_rows=sample_rows,
        eligible_count=eligible_count,
        baseline_column=baseline_column,
        rag_column=rag_column,
    )

    print("=" * 70)
    print("ONE HALLUCINATION EVALUATION FILE CREATED")
    print("=" * 70)
    print("No Ollama, ChromaDB, retrieval, or embeddings were used.")
    print(f"Input questions: {len(rows)}")
    print(f"Questions answered by both systems: {eligible_count}")
    print(f"Sampled paired questions: {len(sample_rows)}")
    print(f"Only output file: {output_path}")
    print("Open it in Excel and fill the two yellow label columns.")
    print("The result table at the top updates automatically.")


if __name__ == "__main__":
    main()